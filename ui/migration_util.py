import openpyxl
import re
from datetime import date, datetime
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from database import (add_category, save_detailed_budget, add_asset, 
                      add_ledger_entry, clear_household_data, get_db_connection)
from PyQt6.QtWidgets import QMessageBox

def to_int(value):
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).replace(",", "").strip()))
    except (TypeError, ValueError):
        return None

def parse_excel_date(value, year, fallback_month=None):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    text = str(value).strip()
    if not text:
        return None

    iso_match = re.search(r"(\d{4})[-./](\d{1,2})[-./](\d{1,2})", text)
    if iso_match:
        y, m, d = map(int, iso_match.groups())
        return f"{y:04d}-{m:02d}-{d:02d}"

    month_day_match = re.search(r"(\d{1,2})\s*[./]\s*(\d{1,2})", text)
    if month_day_match:
        m, d = map(int, month_day_match.groups())
        return f"{year:04d}-{m:02d}-{d:02d}"

    day_match = re.fullmatch(r"\d{1,2}", text)
    if day_match and fallback_month:
        return f"{year:04d}-{fallback_month:02d}-{int(text):02d}"

    return text

def get_import_year(wb):
    if "설정하기" not in wb.sheetnames:
        return datetime.now().year

    sheet = wb["설정하기"]
    for row in sheet.iter_rows(values_only=True):
        for idx, value in enumerate(row):
            if value == "연도 설정":
                for next_value in row[idx + 1:]:
                    year = to_int(next_value)
                    if year:
                        return year
    return datetime.now().year

def find_header_row(sheet):
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [str(value).strip() if value is not None else "" for value in row]
        if "소비날짜" in values and "소득날짜" in values:
            return row_idx, values
    return None, []

def _group_rows(rows, key_indexes):
    grouped = {}
    for row in rows:
        key = tuple(row[index] for index in key_indexes)
        grouped.setdefault(key, []).append(row)
    return grouped

def _write_title_row(sheet, title, max_col):
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell = sheet.cell(1, 1, title)
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4F81BD")
    cell.alignment = Alignment(horizontal="center")

def _write_headers(sheet, row, headers):
    fill = PatternFill("solid", fgColor="5B7DB1")
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row, col, header)
        if header:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

def _set_column_widths(sheet, widths):
    for col, width in widths.items():
        sheet.column_dimensions[get_column_letter(col)].width = width

def _fetch_export_data(hid, year):
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, type, parent_category, sub_category
        FROM categories
        WHERE household_id = ?
        ORDER BY type, parent_category, id
    """, (hid,))
    categories = cursor.fetchall()

    cursor.execute("""
        SELECT month, category_name, amount
        FROM budgets
        WHERE household_id = ? AND year = ?
        ORDER BY month, category_name
    """, (hid, year))
    budgets = cursor.fetchall()

    cursor.execute("""
        SELECT asset_name, initial_balance, current_balance
        FROM assets
        WHERE household_id = ?
        ORDER BY id
    """, (hid,))
    assets = cursor.fetchall()

    cursor.execute("""
        SELECT l.date, l.type, l.amount, l.memo, l.payee, l.payment_method,
               c.parent_category, c.sub_category,
               p.parent_category AS payment_parent, p.sub_category AS payment_name
        FROM ledgers l
        LEFT JOIN categories c ON l.category_id = c.id
        LEFT JOIN categories p ON l.asset_id = p.id
        WHERE l.household_id = ? AND l.date LIKE ?
        ORDER BY l.date ASC, l.id ASC
    """, (hid, f"{year}-%"))
    ledgers = cursor.fetchall()

    conn.close()
    return categories, budgets, assets, ledgers

def _write_settings_sheet(workbook, year, categories):
    sheet = workbook.active
    sheet.title = "설정하기"
    _write_title_row(sheet, "설정하기", 12)
    sheet.cell(2, 1, "연도 설정")
    sheet.cell(2, 3, year)

    type_order = ["소비", "결제수단", "소득", "자본"]
    type_titles = {
        "소비": "소비 분류",
        "결제수단": "결제분류",
        "소득": "소득 분류",
        "자본": "자본, 부채 분류",
    }
    grouped = {}
    for _, category_type, parent, sub in categories:
        grouped.setdefault(category_type, {}).setdefault(parent, []).append(sub)

    row = 4
    for category_type in type_order:
        parents = grouped.get(category_type, {})
        if not parents:
            continue
        section_cell = sheet.cell(row, 1, type_titles[category_type])
        section_cell.font = Font(bold=True, color="FFFFFF")
        section_cell.fill = PatternFill("solid", fgColor="70AD47")
        row += 1
        for parent, subs in parents.items():
            unique_subs = list(dict.fromkeys(subs))
            sheet.cell(row, 1, parent)
            sheet.cell(row, 2, ",".join(unique_subs))
            for col, sub in enumerate(unique_subs, start=3):
                sheet.cell(row, col, sub)
            row += 1
        row += 1

    _set_column_widths(sheet, {1: 18, 2: 36, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16, 8: 16})

def _write_budget_sheet(workbook, year, categories, budgets):
    sheet = workbook.create_sheet("예산 설정")
    expense_parents = []
    for _, category_type, parent, _ in categories:
        if category_type == "소비" and parent not in expense_parents:
            expense_parents.append(parent)
    for _, category_name, _ in budgets:
        if category_name not in expense_parents:
            expense_parents.append(category_name)

    headers = ["총 예산", "월＼대분류"] + expense_parents
    _write_headers(sheet, 1, headers)

    budget_map = {(month, category): amount for month, category, amount in budgets}
    for month in range(1, 13):
        row = month + 2
        sheet.cell(row, 1, f"=SUM(C{row}:{get_column_letter(len(headers))}{row})")
        sheet.cell(row, 2, f"{month}월")
        for col, parent in enumerate(expense_parents, start=3):
            sheet.cell(row, col, budget_map.get((month, parent), 0))
            sheet.cell(row, col).number_format = "#,##0"

    _set_column_widths(sheet, {1: 14, 2: 12, **{col: 16 for col in range(3, len(headers) + 1)}})

def _write_asset_sheet(workbook, assets):
    sheet = workbook.create_sheet("자산 관리")
    _write_title_row(sheet, "자산 관리", 5)
    _write_headers(sheet, 3, ["자산명", "초기 금액", "현재 금액", None, "자본금"])

    capital_amount = 0
    for row_index, (name, initial_balance, current_balance) in enumerate(assets, start=4):
        sheet.cell(row_index, 1, name)
        sheet.cell(row_index, 2, initial_balance or 0)
        sheet.cell(row_index, 3, current_balance or 0)
        sheet.cell(row_index, 2).number_format = "#,##0"
        sheet.cell(row_index, 3).number_format = "#,##0"
        if name == "자본금":
            capital_amount = initial_balance or current_balance or 0

    if not capital_amount:
        capital_amount = sum(initial_balance or 0 for _, initial_balance, _ in assets)

    sheet.cell(3, 6, capital_amount)
    sheet.cell(3, 6).number_format = "#,##0"
    _set_column_widths(sheet, {1: 18, 2: 14, 3: 14, 5: 12, 6: 14})

def _write_month_sheet(workbook, year, month, ledger_rows):
    sheet = workbook.create_sheet(f"{month}월")
    sheet.cell(1, 1, f"{month}월 - 소비 · 소득")
    sheet.cell(1, 3, "월 예산 :")
    sheet.cell(1, 4, f"='예산 설정'!A{month + 2}")
    sheet.cell(1, 5, "현재 지출 :")
    sheet.cell(1, 6, "=SUM(F4:F996)")
    sheet.cell(1, 7, "예산 :")
    sheet.cell(1, 8, '=IF(D1=0,"-",D1-F1)')
    sheet.cell(1, 12, "현재 소득 :")
    sheet.cell(1, 13, "=SUM(M4:M996)")

    headers = [
        "소비날짜", "결제수단", "수단명", "대분류", "항목", "지출금액", "사용처", "코멘트",
        None, "소득날짜", "대분류", "항목", "소득 금액", "소득처", "코멘트"
    ]
    _write_headers(sheet, 3, headers)

    expenses = [row for row in ledger_rows if row[1] == "지출"]
    incomes = [row for row in ledger_rows if row[1] == "수입"]
    max_rows = max(len(expenses), len(incomes), 1)
    for index in range(max_rows):
        row_index = index + 4
        if index < len(expenses):
            date_text, _, amount, memo, payee, payment_method, parent, sub, payment_parent, payment_name = expenses[index]
            sheet.cell(row_index, 1, parse_excel_date(date_text, year, month))
            sheet.cell(row_index, 2, payment_method or payment_parent or "")
            sheet.cell(row_index, 3, payment_name or "")
            sheet.cell(row_index, 4, parent or "")
            sheet.cell(row_index, 5, sub or "")
            sheet.cell(row_index, 6, amount or 0)
            sheet.cell(row_index, 7, payee or "")
            sheet.cell(row_index, 8, memo or "")
            sheet.cell(row_index, 6).number_format = "#,##0"
        if index < len(incomes):
            date_text, _, amount, memo, payee, _, parent, sub, _, _ = incomes[index]
            sheet.cell(row_index, 10, parse_excel_date(date_text, year, month))
            sheet.cell(row_index, 11, parent or "")
            sheet.cell(row_index, 12, sub or "")
            sheet.cell(row_index, 13, amount or 0)
            sheet.cell(row_index, 14, payee or "")
            sheet.cell(row_index, 15, memo or "")
            sheet.cell(row_index, 13).number_format = "#,##0"

    _set_column_widths(sheet, {
        1: 13, 2: 12, 3: 14, 4: 16, 5: 14, 6: 12, 7: 16, 8: 20,
        9: 4, 10: 13, 11: 14, 12: 14, 13: 12, 14: 16, 15: 20
    })

def export_to_excel(hid, year, file_path, parent_widget):
    try:
        categories, budgets, assets, ledgers = _fetch_export_data(hid, year)
        workbook = openpyxl.Workbook()

        _write_settings_sheet(workbook, year, categories)
        _write_budget_sheet(workbook, year, categories, budgets)
        _write_asset_sheet(workbook, assets)

        ledgers_by_month = _group_rows(ledgers, [0])
        month_rows = {month: [] for month in range(1, 13)}
        for rows in ledgers_by_month.values():
            for row in rows:
                ds = parse_excel_date(row[0], year)
                if not ds:
                    continue
                month_rows[int(ds[5:7])].append(row)

        for month in range(1, 13):
            _write_month_sheet(workbook, year, month, month_rows[month])

        workbook.save(file_path)
        QMessageBox.information(parent_widget, "완료", "엑셀 내보내기가 완료되었습니다.")
        return True
    except Exception as e:
        QMessageBox.critical(parent_widget, "오류", f"엑셀 내보내기 중 오류 발생: {e}")
        return False

def import_from_excel(hid, file_path, parent_widget):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        import_year = get_import_year(wb)
        
        reply = QMessageBox.question(
            parent_widget, "데이터 이관 확인",
            "기존 모든 데이터를 삭제하고 엑셀 데이터를 가져오시겠습니까?\n(가져오기 후 프로그램이 최신 상태로 갱신됩니다.)",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.No: return False

        # Clear data for THIS household only
        clear_household_data(hid)

        # 2. Categories
        if '설정하기' in wb.sheetnames:
            sheet = wb['설정하기']
            current_type = "소비"
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if not row[0]: continue
                h = str(row[0]).strip()
                if "소비 분류" in h: current_type = "소비"
                elif "소득 분류" in h: current_type = "소득"
                elif "결제분류" in h: current_type = "결제수단"
                elif "자본" in h and "부채" in h: current_type = "자본"
                if row[2]:
                    parent = row[0]
                    if parent in ["연도 설정", "소비 분류", "결제분류", "소득 분류", "자본, 부채 분류"]: continue
                    for col_idx in range(2, len(row)):
                        sub = row[col_idx]
                        if sub: add_category(hid, current_type, parent, sub)

        # 3. Budget
        if '예산 설정' in wb.sheetnames:
            sheet = wb['예산 설정']
            headers = [c for c in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)) if c]
            for row in sheet.iter_rows(min_row=3, values_only=True):
                ms = str(row[1]) if row[1] else ""
                if "월" in ms:
                    month = int(ms.replace("월", ""))
                    for i, amt in enumerate(row[2:]):
                        if i + 2 < len(headers) and amt:
                            save_detailed_budget(hid, import_year, month, headers[i+2], int(amt))

        # 4. Assets
        if '자산 관리' in wb.sheetnames:
            sheet = wb['자산 관리']
            for row in sheet.iter_rows(min_row=1, values_only=True):
                values = list(row)
                for idx, value in enumerate(values):
                    if value == "자본금":
                        for next_value in values[idx + 1:]:
                            amount = to_int(next_value)
                            if amount is not None:
                                add_asset(hid, "자본금", amount)
                                break
                        break

        # 5. Ledger
        cat_map = {}
        conn = get_db_connection(); cursor = conn.cursor()
        cursor.execute("SELECT id, type, parent_category, sub_category FROM categories WHERE household_id = ?", (hid,))
        for r in cursor.fetchall(): cat_map[(r[1], r[2], r[3])] = r[0]
        conn.close()

        for month in range(1, 13):
            sheet_name = f"{month}월"
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                header_row, headers = find_header_row(sheet)
                if not header_row:
                    continue
                expense_start = headers.index("소비날짜")
                income_start = headers.index("소득날짜")

                for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                    expense_amount = to_int(row[expense_start + 5] if len(row) > expense_start + 5 else None)
                    if len(row) > expense_start and row[expense_start] and expense_amount:
                        ds = parse_excel_date(row[expense_start], import_year, month)
                        pm = str(row[expense_start + 1]).strip() if len(row) > expense_start + 1 and row[expense_start + 1] else ""
                        an = str(row[expense_start + 2]).strip() if len(row) > expense_start + 2 and row[expense_start + 2] else ""
                        p = str(row[expense_start + 3]).strip() if len(row) > expense_start + 3 and row[expense_start + 3] else ""
                        s = str(row[expense_start + 4]).strip() if len(row) > expense_start + 4 and row[expense_start + 4] else ""
                        payee = str(row[expense_start + 6]).strip() if len(row) > expense_start + 6 and row[expense_start + 6] else ""
                        memo = str(row[expense_start + 7]).strip() if len(row) > expense_start + 7 and row[expense_start + 7] else ""
                        cat_id = cat_map.get(("소비", p, s))
                        asset_id = cat_map.get(("결제수단", pm, an))
                        add_ledger_entry(hid, ds, "지출", cat_id, asset_id, expense_amount, memo, payee, pm)

                    income_amount = to_int(row[income_start + 3] if len(row) > income_start + 3 else None)
                    if len(row) > income_start and row[income_start] and income_amount:
                        ds = parse_excel_date(row[income_start], import_year, month)
                        p = str(row[income_start + 1]).strip() if len(row) > income_start + 1 and row[income_start + 1] else ""
                        s = str(row[income_start + 2]).strip() if len(row) > income_start + 2 and row[income_start + 2] else ""
                        payee = str(row[income_start + 4]).strip() if len(row) > income_start + 4 and row[income_start + 4] else ""
                        memo = str(row[income_start + 5]).strip() if len(row) > income_start + 5 and row[income_start + 5] else ""
                        cat_id = cat_map.get(("소득", p, s))
                        add_ledger_entry(hid, ds, "수입", cat_id, None, income_amount, memo, payee, "")

        QMessageBox.information(parent_widget, "완료", "엑셀 데이터 이관이 성공적으로 완료되었습니다.")
        return True
    except Exception as e:
        QMessageBox.critical(parent_widget, "오류", f"엑셀 임포트 중 오류 발생: {e}"); return False
